import openpyxl
from PyQt6.QtWidgets import QMessageBox

def export_to_excel(data_table):
    if data_table.item(0, 0) is None:
        QMessageBox.warning(None, "No Data", "There is no data to export to Excel.")
        return

    file_path = "F:\\Project\\GUI\\Database_test.xlsx"

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        next_row = sheet.max_row + 2

        for row in range(data_table.rowCount()):
            if data_table.item(row, 0) is None:
                break
            for col in range(data_table.columnCount()):
                item = data_table.item(row, col)
                if item is not None:
                    sheet.cell(row=next_row + row, column=col + 2, value=item.text())

        workbook.save(file_path)
        print(f"Data appended to {file_path}")

    except PermissionError:
        QMessageBox.warning(None, "File Open Error", "Please close the Excel file before uploading.")
    except Exception as e:
        QMessageBox.warning(None, "Error", f"An error occurred: {e}")
